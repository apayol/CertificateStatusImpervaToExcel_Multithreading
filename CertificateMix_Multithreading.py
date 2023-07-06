#!/usr/bin/env python

################################################################################
### Nombre del script: CertificateMix_Multithreading.py
### Autor: Adrián Payol Montero
### Equipo de trabajo: WAF Telefónica
### Versión: 1.0
### Explicación: Salida a excels individuales por entidad con 2 hojas y optimizado con multi-hilos
### Estado de los certificados de cada uno de los dominios y próximas acciones.
################################################################################

import requests
import warnings
import datetime
import json
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.styles import Alignment
import threading

#Credenciales y headers
IMPERVA_API_ID = "xxx"
IMPERVA_API_KEY = "yyyyyy"

IMPERVA_headers = {
   "x-API-key": IMPERVA_API_KEY,
   "x-API-id": IMPERVA_API_ID,
   "Content-Type" : "application/json"
}

def getAccountName(account_id):

   url = "https://my.imperva.com/api/prov/v1/account?account_id=" + str(account_id)
   try:
      response = requests.post(url, headers=IMPERVA_headers)
      jsonresp = json.dumps(response.json())
      response_dict = json.loads(jsonresp)
      return response_dict["account"]["account_name"]
   except:
      print("API interaction error: GetAccountName" + str(response))

def getAccountIdList():

    generalAccountId = "1281572"
    accountIdList = []
    url = "https://my.imperva.com/api/prov/v1/accounts/listSubAccounts?account_id=" + str(generalAccountId) + "&page_size=100&page_num=0"
    try:
        response = requests.post(url, headers=IMPERVA_headers)
        jsonresp = json.dumps(response.json())
        response_dict = json.loads(jsonresp)

        for i in range(len(response_dict["resultList"])):
            accountIdList.append ((response_dict["resultList"][i]["sub_account_id"]))
        return accountIdList

    except:
        print("API interaction error: GetAccountName" + str(response))

def getCustomCertData (siteId):
    url = "https://api.imperva.com/certificates-ui/v3/certificates?extSiteId=" + str(siteId) + "&certType=CUSTOM_CERT"
    response = requests.get(url, headers=IMPERVA_headers, data="", verify=False)
    jsonresp = json.dumps(response.json())
    response_dict = json.loads(jsonresp)
    try:
        if response_dict["data"][0]['status'] == "ACTIVE":

            customCertActive = "Active"
            expDateCustomCertRAW = response_dict["data"][0]['expirationDate']
            expDateCustomCert = datetime.datetime.fromtimestamp(int(expDateCustomCertRAW / 1000)).strftime(
                '%d/%m/%Y %H:%M:%S')
        elif response_dict["data"][0]['status'] == "NEAR_EXPIRATION":
            customCertActive = "Active"
            expDateCustomCertRAW = response_dict["data"][0]['expirationDate']
            expDateCustomCert = datetime.datetime.fromtimestamp(int(expDateCustomCertRAW / 1000)).strftime(
                '%d/%m/%Y %H:%M:%S')
        elif response_dict["data"][0]['status'] == "EXPIRED":
            customCertActive = "Expired"
            expDateCustomCertRAW = response_dict["data"][0]['expirationDate']
            expDateCustomCert = datetime.datetime.fromtimestamp(int(expDateCustomCertRAW / 1000)).strftime(
                '%d/%m/%Y %H:%M:%S')
        else:
            customCertActive = "Not active"
            expDateCustomCert = "-"

    except:
        # Vacía la salida a la llamada a CUSTOM_CERT, no lo tiene añadido.
        customCertActive = "Not active"
        expDateCustomCert = "-"

    return customCertActive, expDateCustomCert

def getImpervaCertificate(siteId):

    with requests.Session() as s:
        url = "https://api.imperva.com/certificates-ui/v3/certificates?extSiteId=" + str(siteId) + "&certType=ATLAS"
        response = requests.get(url, headers=IMPERVA_headers, data="", verify=False)
        jsonresp = json.dumps(response.json())
        response_dict = json.loads(jsonresp)

        if response.status_code == 200:
            jsonresp = response.json()
            if not jsonresp["data"]: #Si data está vacío, devolvemos None
                return None
        return jsonresp #Si data tiene algo, devolvemos el contenido

def getStatusImpervaCertificate(data):

    status=""
    if data is not None:
        if "data" in data:
            for d in data["data"]:
                if "sans" in d:
                    for s in d["sans"]:
                        if not s["expirationDate"]:
                            status="TXT must be added"
                            break
                        else:
                            if s["status"] == "PENDING_USER_ACTION":
                                status="Revalidation Required"
                                break
                            elif s["status"] == "VALIDATED":
                                status="Published"
                                break
            return status
    else:
        # Caso Cert Imperva sin activar
        return "-"

    return status

def getexpirationDateGlobalsignImpervaCertificate(siteId, statusImpervaCertificate, impervaCertificate):
    url = "https://api.imperva.com/certificates-ui/v3/certificates?extSiteId=" + str(siteId) + "&certType=ATLAS"
    response = requests.get(url, headers=IMPERVA_headers, data="", verify=False)
    jsonresp = json.dumps(response.json())
    response_dict = json.loads(jsonresp)

    try:
        if statusImpervaCertificate == "TXT must be added":
            expDateImpervaCert = "-"
        if statusImpervaCertificate == "Revalidation Required":
            # Mail revalidation date
            expDateImpervaCert = response_dict["data"][0]['expirationDate'] # no es la fecha!
            expDateMinus7Days = expDateImpervaCert - 604800000
            expDateImpervaCert = datetime.datetime.fromtimestamp(int(expDateMinus7Days / 1000)).strftime(
                '%d/%m/%Y %H:%M:%S')
        if statusImpervaCertificate == "Published":
            expDateImpervaCert = response_dict["data"][0]['sans'][0]['expirationDate']
            expDateImpervaCert = datetime.datetime.fromtimestamp(int(expDateImpervaCert / 1000)).strftime('%d/%m/%Y %H:%M:%S')
        if impervaCertificate == "Not Activated":
            expDateImpervaCert = "-"

        return expDateImpervaCert
    except:
        expDateImpervaCert = "getImpervaCertExpDate Error"
        return expDateImpervaCert

#sin uso
def getHost (data, siteId, getImpervaCertificate):
    url = "https://api.imperva.com/certificates-ui/v3/certificates?extSiteId=" + str(siteId) + "&certType=ATLAS"
    response = requests.get(url, headers=IMPERVA_headers, data="", verify=False)
    jsonresp = json.dumps(response.json())
    response_dict = json.loads(jsonresp)

    status=""
    host=""
    if getImpervaCertificate == "Not Activated":
        host = "-"
    if data is not None:
        if "data" in data:
            for d in data["data"]:
                if "sans" in d:
                    for s in d["sans"]:
                        if s["approverFqdn"] is not None:
                            host = s["approverFqdn"]
                            break

    return host

#sin uso
def getTXT (data, siteId, getImpervaCertificate):
    url = "https://api.imperva.com/certificates-ui/v3/certificates?extSiteId=" + str(siteId) + "&certType=ATLAS"
    response = requests.get(url, headers=IMPERVA_headers, data="", verify=False)
    jsonresp = json.dumps(response.json())
    response_dict = json.loads(jsonresp)

    txt=""
    if getImpervaCertificate == "Not Activated":
        txt = "-"
    if data is not None:
        if "data" in data:
            for d in data["data"]:
                if "sans" in d:
                    for s in d["sans"]:
                        if s["verificationCode"] is not None:
                            txt = s["verificationCode"]
                            break

    return txt

#sin uso
def getExpTxtDate (siteId):
    url = "https://api.imperva.com/certificates/v3/instructions?extSiteId=" + str(siteId) + "&validationMethod=DNS&certificateType=ATLAS"

    try:
        response = requests.get(url, headers=IMPERVA_headers, data="", verify=False)
        jsonresp = json.dumps(response.json())
        response_dict = json.loads(jsonresp)

        expTxtDateRAW = response_dict["data"][0]["verificationCodeExpirationDate"]
        expTxtDate = datetime.datetime.fromtimestamp(int(expTxtDateRAW / 1000)).strftime('%d/%m/%Y %H:%M:%S')

    except:
        expTxtDate = "-"

    return expTxtDate

def getCnameRevalidationStatus (siteId, accountId, siteName):
    url = "https://api.imperva.com/certificates-ui/v3/account/ssl-settings?extSiteId=" + str(siteId) + "&caid=" + str(accountId)
    response = requests.get(url, headers=IMPERVA_headers, data="", verify=False)
    jsonresp = json.dumps(response.json())
    response_dict = json.loads(jsonresp)
    CNAMERevalidationStatus = "-"

    domainNameStripped = siteName.split(".", )
    domainNameStripped2 = siteName.split(".", )

    arrayLength = len(domainNameStripped)
    domainNameStripped = domainNameStripped[arrayLength - 2] + "." + domainNameStripped[arrayLength - 1]
    domainNameStripped2 = domainNameStripped2[arrayLength - 3] + "." + domainNameStripped2[arrayLength - 2] + "." + \
                          domainNameStripped2[arrayLength - 1]

    # si array vacío, len es 0
    if len(response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"]) == 0:
        CNAMERevalidationStatus = "-"
    else:
        for i in range(len(response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"])):
            if response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"][i]["name"] == siteName:
                CNAMERevalidationStatus = response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"][i]["status"]
                break
            elif response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"][i]["name"] == domainNameStripped:
                CNAMERevalidationStatus = response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"][i]["status"]
            elif response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"][i]["name"] == domainNameStripped2:
                CNAMERevalidationStatus = response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"][i]["status"]

    return CNAMERevalidationStatus

def getDnsRecordDomain (siteId, accountId, siteName):
    url = "https://api.imperva.com/certificates-ui/v3/account/ssl-settings?extSiteId=" + str(siteId) + "&caid=" + str(accountId)
    response = requests.get(url, headers=IMPERVA_headers, data="", verify=False)
    jsonresp = json.dumps(response.json())
    response_dict = json.loads(jsonresp)

    DNSRecordDomain = "-"
    domainNameStripped = siteName.split(".", )
    domainNameStripped2 = siteName.split(".", )

    arrayLength = len(domainNameStripped)
    domainNameStripped = domainNameStripped[arrayLength - 2] + "." + domainNameStripped[arrayLength - 1]
    domainNameStripped2 = domainNameStripped2[arrayLength - 3] + "." + domainNameStripped2[arrayLength - 2] + "." + domainNameStripped2[arrayLength - 1]

    # si array vacío, len es 0
    if len(response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"]) == 0:
        DNSRecordDomain = "-"
    else:
        for i in range(len(response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"])):
            if response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"][i]["name"] == siteName:
                DNSRecordDomain = response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"][i]["cnameRecordHost"]
                break
            elif response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"][i]["name"] == domainNameStripped:
                DNSRecordDomain = response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"][i]["cnameRecordHost"]
            elif response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"][i]["name"] == domainNameStripped2:
                DNSRecordDomain = response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"][i]["cnameRecordHost"]
    return DNSRecordDomain

def getCnameRecord(siteId, accountId, siteName):
    url = "https://api.imperva.com/certificates-ui/v3/account/ssl-settings?extSiteId=" + str(siteId) + "&caid=" + str(accountId)
    response = requests.get(url, headers=IMPERVA_headers, data="", verify=False)
    jsonresp = json.dumps(response.json())
    response_dict = json.loads(jsonresp)

    CNAMERecord = "-"
    domainNameStripped = siteName.split(".", )
    domainNameStripped2 = siteName.split(".", )

    arrayLength = len(domainNameStripped)
    domainNameStripped = domainNameStripped[arrayLength - 2] + "." + domainNameStripped[arrayLength - 1]
    domainNameStripped2 = domainNameStripped2[arrayLength - 3] + "." + domainNameStripped2[arrayLength - 2] + "." + \
                          domainNameStripped2[arrayLength - 1]

    # si array vacío, len es 0
    if len(response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"]) == 0:
        CNAMERecord = "-"
    else:
        for i in range(len(response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"])):
            if response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"][i]["name"] == siteName:
                CNAMERecord = response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"][i]["cnameRecordValue"]
                break
            elif response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"][i]["name"] == domainNameStripped:
                CNAMERecord = response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"][i]["cnameRecordValue"]
            elif response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"][i]["name"] == domainNameStripped2:
                CNAMERecord = response_dict["data"][0]["impervaCertificate"]["delegation"]["allowedDomainsForCNAMEValidation"][i]["cnameRecordValue"]

    return CNAMERecord

def storageMatrix(account_id):
   # Paginar los sites de la cuenta
   page_id = 0
   count_item_per_page = 20
   finished_pagination = False
   dataMatrix = []
   count = 1

   while not finished_pagination:
      url = "https://my.imperva.com/api/prov/v1/sites/list?page_size=" + str(count_item_per_page) + "&page_num=" + str(
         page_id) + "&account_id=" + str(account_id)
      response = requests.post(url, headers=IMPERVA_headers)
      jsonresp = json.dumps(response.json())
      response_dict = json.loads(jsonresp)

      if not response_dict["sites"]:
         finished_pagination = True
         break

      for i in range(len(response_dict["sites"])):
         try:
            siteId = response_dict["sites"][i]["site_id"]
            accountId = response_dict["sites"][i]["account_id"]
            # dataMatrix[1]
            siteName = response_dict["sites"][i]["domain"]
            # dataMatrix[0]
            accountName = getAccountName(response_dict["sites"][i]["account_id"])
            # dataMatrix[2,3]
            customCertificate, expirationDateCustomCertificate = getCustomCertData(siteId)
            # dataMatrix[4]
            impervaCertificate = getImpervaCertificate(siteId)
            if impervaCertificate is None:
                impervaCertificate = "Not Activated"
            else:
                impervaCertificate = "Active"
            # dataMatrix[5]
            statusImpervaCertificate = getStatusImpervaCertificate(getImpervaCertificate(siteId))
            # dataMatrix[6]
            expirationDateGlobalsignImpervaCertificate = getexpirationDateGlobalsignImpervaCertificate(siteId, statusImpervaCertificate, impervaCertificate)
            # dataMatrix[7]
            cnameRevalidationStatus = getCnameRevalidationStatus (siteId, accountId, siteName)
            # dataMatrix[8]
            dnsRecordDomain = getDnsRecordDomain (siteId, accountId, siteName)
            # dataMatrix[9]
            cnameRecord = getCnameRecord(siteId, accountId, siteName)

            # Almacenamos en lista de túplas
            dataMatrix.append((accountName, siteName, customCertificate, expirationDateCustomCertificate,
                               impervaCertificate, statusImpervaCertificate, expirationDateGlobalsignImpervaCertificate,
                               cnameRevalidationStatus, dnsRecordDomain, cnameRecord))

            # Trazabilidad de la salida
            print(str(count) + " - " + accountName + " - " + siteName)
            count += 1

         except:
            print("Data storage failure: " + str(siteId) + " | " + str(count) + " - " + str(accountName) + " - " + str(siteName))
            count += 1
            raise

      page_id += 1

   return dataMatrix

def getUniqueEntities(dataMatrix):
   uniqueEntitiesList = []
   #Recorremos la columna de entidad de la matriz de datos
   for i in range(len(dataMatrix)):
      accountName = dataMatrix[i][0]
      # Almacenamos en lista de túplas todas las entidades
      uniqueEntitiesList.append((accountName))

   #Nos quedamos con los valores únicos de la lista
   uniqueEntitiesList = set(uniqueEntitiesList)
   #Convertimos de tipo set a arraylist
   uniqueEntitiesList = list(uniqueEntitiesList)

   return uniqueEntitiesList

def createExcel(accountName):
   # Creamos excel
   wb = openpyxl.Workbook()
   today = datetime.date.today()
   excelName = "CertificateStatus-" + str(accountName) + "-" + str(today) + ".xlsx"

   # Creamos hojas del excel (primero la derecha)
   ws2 = wb.active
   ws2.title = "DelegatedCertificateStatus"
   ws1 = wb.create_sheet("CertificateStatus", 0)

   # Cabeceras hoja 1
   ws1['A1'] = "Entity"
   ws1['B1'] = "Site"
   ws1['C1'] = "Custom Certificate"
   ws1['D1'] = "Expiration Date Custom Certificate"
   ws1['E1'] = "Imperva Certificate"
   ws1['F1'] = "Status Imperva Certificate"
   ws1['G1'] = "Expiration Date Globalsign Imperva Certificate"

   # Cabeceras hoja 2
   ws2['A1'] = "Entity"
   ws2['B1'] = "Site"
   ws2['C1'] = "CNAME Revalidation Status"
   ws2['D1'] = "DNS Record Domain"
   ws2['E1'] = "CNAME Record"

   # Cerramos
   wb.save(excelName)

def printMatrix(dataMatrix):
   today = datetime.date.today()

   #Recorremos la matriz de datos
   for i in range(len(dataMatrix)):
      accountName = dataMatrix[i][0]
      # Abrimos excel
      excelName = "CertificateStatus-" + str(accountName) + "-" + str(today) + ".xlsx"
      wb = load_workbook(excelName)

      # Abrimos e imprimimos hoja 1
      ws1 = wb["CertificateStatus"]
      lastRow = ws1.max_row
      ws1['A' + str(lastRow + 1)] = dataMatrix[i][0]
      ws1['B' + str(lastRow + 1)] = dataMatrix[i][1]
      ws1['C' + str(lastRow + 1)] = dataMatrix[i][2]
      ws1['D' + str(lastRow + 1)] = dataMatrix[i][3]
      ws1['E' + str(lastRow + 1)] = dataMatrix[i][4]
      ws1['F' + str(lastRow + 1)] = dataMatrix[i][5]
      ws1['G' + str(lastRow + 1)] = dataMatrix[i][6]

      # Abrimos e imprimimos hoja 2
      ws2 = wb["DelegatedCertificateStatus"]
      lastRow = ws2.max_row
      ws2['A' + str(lastRow + 1)] = dataMatrix[i][0]
      ws2['B' + str(lastRow + 1)] = dataMatrix[i][1]
      ws2['C' + str(lastRow + 1)] = dataMatrix[i][7]
      ws2['D' + str(lastRow + 1)] = dataMatrix[i][8]
      ws2['E' + str(lastRow + 1)] = dataMatrix[i][9]

      # Damos formato a los excels #
      headersFill = PatternFill("solid", start_color="c4d4e9")
      headersFont = Font (color="060f14", bold = True)
      horizontalAlignment = Alignment(horizontal='center')

      # 1.Formato a cabeceras
      for cell in ws1["1:1"]:
          cell.font = headersFont
          cell.fill = headersFill
      for cell in ws2["1:1"]:
          cell.font = headersFont
          cell.fill = headersFill

      # 2.Ancho de columnas
      for column_cells in ws1.columns:
          length = max(len(str(cell.value)) for cell in column_cells)
          ws1.column_dimensions[column_cells[0].column_letter].width = length
      for column_cells in ws2.columns:
          length = max(len(str(cell.value)) for cell in column_cells)
          ws2.column_dimensions[column_cells[0].column_letter].width = length

      # 3.Centrar texto horizontalmente
      for row in range(1, ws1.max_row + 1):
          for col in range(1, ws1.max_column + 1):
              cell = ws1.cell(row, col)
              cell.alignment = horizontalAlignment
      for row in range(1, ws2.max_row + 1):
          for col in range(1, ws2.max_column + 1):
              cell = ws2.cell(row, col)
              cell.alignment = horizontalAlignment

      #Cerramos excel
      wb.save(excelName)

def main(account_id):
   # Ignorar warnings
   warnings.filterwarnings("ignore")

   print("### CERTIFICATE STATUS | ACCOUNT ID: " + str(account_id) + " ###")

   #Almacenamos en dataMatrix la matriz con los datos
   dataMatrix = storageMatrix(account_id)

   #Sacamos la lista con las entidades únicas
   uniqueEntities = getUniqueEntities (dataMatrix)

   #Creamos los excels con páginas y cabeceras
   for i in range(len(uniqueEntities)):
      accountName = uniqueEntities[i]
      createExcel(accountName)

   #Imprimimos los datos en sus excels y damos formato
   printMatrix(dataMatrix)

if __name__ == '__main__':
    # Imprimo hora inicial
    now = datetime.datetime.now()
    now = now.replace(microsecond=0)
    print(f"Inicio de la ejecución: {now}")

    # EJECUCIÓN MULTIHILO #

    accounts_ids = getAccountIdList()
    h={} # Creo diccionario vacío

    # Declaro los hilos (uno por subcuenta)
    for i in range(len(accounts_ids)):
        # En cada h0,h1... creo su hilo
        h["h{0}".format(i)] = (threading.Thread(target=lambda: main(accounts_ids[i])))

    # Ejecuto
    for i in range(len(accounts_ids)):
        h["h{0}".format(i)].start()

    # Cierro hilos
    for i in range(len(accounts_ids)):
        h["h{0}".format(i)].join()

    # Los hilos terminaron de ejecutar y acabó el programa

    # Imprimo hora final
    now = datetime.datetime.now()
    now = now.replace(microsecond=0)
    print(f"Fin de la ejecución: {now}")
